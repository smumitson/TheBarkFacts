"""Print a sample of existing profiles to match style."""
import openpyxl
wb = openpyxl.load_workbook('dog_facts_7.3.2026.xlsx')
ws = wb['deep_dive_profiles']
headers = [ws.cell(1, c).value for c in range(1, 15)]

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] in ('Borzoi', 'Whippet', 'Basenji'):
        print(f'=== {row[0]} ===')
        for h, v in zip(headers, row):
            print(f'  {h}: {repr(v)}')
        print()

"""Read deep_dive_profiles for batch 5 breeds and print all fields."""
import openpyxl

wb = openpyxl.load_workbook('dog_facts_7.3.2026.xlsx')
ws = wb['deep_dive_profiles']
headers = [ws.cell(1, c).value for c in range(1, 15)]
print('Headers:', headers)
print()

SLUGS = ['saluki', 'samoyed', 'shiba inu', 'siberian husky', 'skye terrier', 'vizsla', 'weimaraner', 'whippet']

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] and row[0].lower() in SLUGS:
        print(f'=== {row[0]} ===')
        for h, v in zip(headers, row):
            print(f'  {h}: {repr(v)}')
        print()
